import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.formatting.rule import ColorScaleRule
from mpl_toolkits.axes_grid1 import make_axes_locatable

def list_material(material_dir="Materials_data"):
    material_files = []
    if os.path.isdir(material_dir):
        material_files = [
            f for f in os.listdir(material_dir) if f.endswith(".txt")
        ]
    return material_files

def convert_to_number(value):
    """
    嘗試將字串依序轉換為 int、float 或 complex，
    如果無法轉換則保留為 str
    """
    for converter in (int, float, complex):
        try:
            return converter(value)
        except ValueError:
            pass
    return value

def make_colorbar_with_padding(ax):
    """
    Create colorbar axis that fits the size of a plot - detailed here: http://chris35wills.github.io/matplotlib_axis/
    """
    divider = make_axes_locatable(ax)
    cax = divider.append_axes("right", size="5%", pad=0.1)
    return(cax)

def add_scatter_plot(file_name, data_length):
    # Load the workbook and the active sheet
    workbook = load_workbook(file_name)
    sheet = workbook["Orders"]

    # Create the scatter chart
    chart = ScatterChart()
    chart.title = "Target Orders Scatter Plot"
    chart.x_axis.title = "X Target Order"
    chart.y_axis.title = "Y Target Order"
    chart.style = 2  # Default scatter chart style

    # Define data for the chart
    x_values = Reference(sheet, min_col=1, min_row=2, max_row=data_length + 1)
    y_values = Reference(sheet, min_col=2, min_row=2, max_row=data_length + 1)
    series = Series(y_values, x_values, title="Orders")
    series.marker.symbol = "circle"  # Use circles for points
    series.marker.size = 7  # Set marker size
    series.graphicalProperties.line.noFill = True  # Remove line between points

    # Add series to chart
    chart.series.append(series)

    # Position the chart
    sheet.add_chart(chart, "D5")

    # Save the workbook
    workbook.save(file_name)

def read_excel_orders(file_name, sheet_name="Orders"):
    # 讀取 Excel 檔案中的指定工作表
    df = pd.read_excel(file_name, sheet_name=sheet_name)
    
    # 檢查是否包含所需的欄位
    if "X Target Order" not in df.columns or "Y Target Order" not in df.columns:
        raise ValueError("Excel file does not contain required columns: 'X Target Order' and 'Y Target Order'")

    # 提取 X 和 Y Target Order 並轉為列表
    xtar_order_list = df["X Target Order"].tolist()
    ytar_order_list = df["Y Target Order"].tolist()

    return xtar_order_list, ytar_order_list

def save_diffraction_intensity_to_excel(Diffracive_Intensity, filename="Diffracive_Intensity_with_axes.xlsx"):
    """
    Save an N x N numpy array to an Excel file with X and Y axes and apply color scale formatting.

    Parameters:
        Diffracive_Intensity (np.ndarray): N x N numpy array containing intensity values.
        filename (str): Name of the Excel file to save. Default is 'Diffracive_Intensity_with_axes.xlsx'.
    """
    Nx = Diffracive_Intensity.shape[0]
    Ny = Diffracive_Intensity.shape[1]
    """ if Diffracive_Intensity.shape[0] != Diffracive_Intensity.shape[1]:
        raise ValueError("Diffracive_Intensity must be a square matrix (N x N).") """

    # Create Excel workbook and sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Intensity Data"

    # Calculate center for axes
    cx = Nx // 2
    cy = Ny // 2

    # Write X-axis labels (1st row, starting from 2nd column)
    for j in range(1, Ny + 1):
        sheet.cell(row=1, column=j + 1, value=j -1 - cy)

    # Write Y-axis labels (1st column, starting from 2nd row)
    for i in range(1, Nx + 1):
        sheet.cell(row=i + 1, column=1, value=i -1 - cx)

    # Write intensity data (starting from B2)
    for i in range(Nx):
        for j in range(Ny):
            sheet.cell(row=i + 2, column=j + 2, value=Diffracive_Intensity[i, j])

    # Apply color scale formatting
    color_scale_rule = ColorScaleRule(
        start_type="min", start_color="FFFF00",  # Min value: Red
        mid_type="percentile", mid_value=5, mid_color="FFFF00",  # Mid value: Yellow
        end_type="max", end_color="00FF00"  # Max value: Green
    )

    # Define data range dynamically and apply formatting
    range_start = "B2"
    range_end = "TT1069"
    sheet.conditional_formatting.add(f"{range_start}:{range_end}", color_scale_rule)

    # Save the workbook
    workbook.save(filename)
    print(f"Excel file save as :'{filename}'")

def createfolder(save_dir, folder_name):
    base_dir = os.path.join(save_dir, folder_name)  # 基礎資料夾名稱
    new_dir = base_dir
    count = 1
    # 如果資料夾已存在，則在名稱後面加上 (數字)
    while os.path.exists(new_dir):
        new_dir = f"{base_dir}({count})"
        count += 1
    os.makedirs(new_dir)
    return new_dir

def split_image_to_layers(img, num_layers=5):
    """
    將灰階圖像分割成多個層次
    
    參數:
        img: numpy.ndarray, 輸入的灰階圖像
        num_layers: int, 要分割的層數
        
    返回:
        list: 包含所有層的列表
    """
    gray_step = 255 / num_layers
    # 創建每層的遮罩
    layers_mask = [
        (img > (gray_step * i - 1)) * (img < (gray_step * i + 1))
        for i in range(1, num_layers)
    ]
    # 疊加層次
    return [sum(layers_mask[i:]) for i in range(len(layers_mask))]